from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from dependencies import get_db
from models import Product
from schemas import ProductRead
from openpyxl import load_workbook
from auth_utils import get_current_user

router = APIRouter()

@router.get("/products", response_model=list[ProductRead])
def get_product(db: Session = Depends(get_db)):
    return db.query(Product).all()

@router.get("/products/{id}", response_model=ProductRead)
def get_id_product(id: int,db: Session = Depends(get_db)):
    db_product = db.query(Product).filter(Product.id == id).first()

    if not db_product:
        raise HTTPException(
            status_code=404, 
            detail="Товар не найден"
        )

    return db_product

@router.post("/products/import")
def import_products(file: UploadFile = File(...),db: Session = Depends(get_db),current_user: str = Depends(get_current_user)):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(
            status_code=400,
            detail="Нужен файл Excel"
        )

    wb = load_workbook(file.file)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=6, values_only=True))



    created = 0
    errors = []
    for index, row in enumerate(rows, start=6):
            brand, name, price, old_price, category, in_stock, image = row

            if not brand and not name:
                continue

            if not brand or not name:
                errors.append(f"Строка {index}: нет бренда или названия")
                continue
            if not price or not isinstance(price, (int, float)):
                errors.append(f"Строка {index}: цена должна быть числом")
                continue
            brand = brand.strip()
            name = name.strip()

            in_stock_value = True
            if in_stock and str(in_stock).strip().lower() == "нет":
                in_stock_value = False

            db_product = Product(
            brand=brand,
            name=name,
            price=int(price),
            old_price=int(old_price) if old_price else None,
            category=category,
            in_stock=in_stock_value,
            image=image
            )

            db.add(db_product)
            created += 1
    db.commit()

    return {
    "created": created,
    "errors_count": len(errors),
    "errors": errors
}